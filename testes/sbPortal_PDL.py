from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl


def sbPortal_PDL():
    # Carregar a planilha de dados
    workbook = openpyxl.load_workbook('caminho_para_sua_planilha.xlsx')
    sheet = workbook['Dados']

    # Inicializar o WebDriver (Chrome)
    driver_path = r'C:\caminho\para\seu\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path)
    wait = WebDriverWait(driver, 10)

    try:
        # Inicia o navegador e navega para a URL
        driver.get("URL_DA_PAGINA")
        driver.maximize_window()

        # Preenche o campo de login com o nome do usuário
        username = "SeuNomeDeUsuario"  # Você pode obter isso dinamicamente, se necessário
        driver.find_element(By.XPATH, "ELEMENT_PATH").send_keys(username)
        driver.find_element(By.ID, "botaoLogin").click()

        # Aguardar até que o código de verificação seja inserido manualmente pelo usuário
        input("Entre no site com o código de verificação enviado por email e pressione Enter para continuar...")

        # Aguarda a presença do elemento "Menu" para continuar
        wait.until(EC.presence_of_element_located((By.ID, "Menu")))

        # Encontrar e clicar na carteira especificada
        carteira = driver.find_element(By.XPATH, "ELEMENT_PATH")
        carteira.click()
        time.sleep(0.5)

        # Digita "PDL" no campo de menu e clica para entrar
        driver.find_element(By.ID, "Menu").send_keys("PDL")
        driver.find_element(By.ID, "entrarMenu").click()
        time.sleep(5)

        # Inicia o processamento dos contratos
        l = 2
        while sheet.cell(row=l, column=2).value is not None:
            nr_contrato_planilha = sheet.cell(row=l, column=2).value

            # Preenche o campo de contrato
            contrato_field = driver.find_element(By.ID, "numero-contrato")
            contrato_field.clear()
            contrato_field.send_keys(nr_contrato_planilha)
            time.sleep(5)

            # Preenche o tipo de liquidação
            tipo_liq_web = driver.find_element(By.NAME, "CodigoTipoLiquidacao")
            option_tipo_liq_web = tipo_liq_web.find_element(By.XPATH, "//option[text()='Liquidação Por Portabilidade']")
            option_tipo_liq_web.click()
            time.sleep(5)

            # Loop para cada data de liquidação
            for i in range(8, 13, 2):
                driver.refresh()
                time.sleep(5)
                data_liq_planilha = sheet.cell(row=l, column=i).value

                # Preenche a data de liquidação
                dt_liq_web = driver.find_element(By.NAME, "DataReferencia")
                dt_liq_web.clear()
                dt_liq_web.send_keys(data_liq_planilha)
                time.sleep(5)

                # Clica no botão de consulta
                botao_consulta = driver.find_element(By.ID, "opcCON")
                botao_consulta.click()
                time.sleep(5)

                # Aguarda até que o valor seja carregado
                wait.until(EC.presence_of_element_located((By.ID, "valorDividaLiquidacao")))
                valorDivida = driver.find_element(By.ID, "valorDividaLiquidacao").get_attribute("value")

                # Exibe o valor capturado e armazena na planilha
                print(f"Valor da dívida: {valorDivida}")
                sheet.cell(row=l, column=i + 1).value = valorDivida
                time.sleep(5)

            # Incrementa a linha para o próximo contrato
            l += 1
            print(f"Próximo contrato na linha: {l}")

        # Salva as alterações na planilha
        workbook.save('caminho_para_sua_planilha.xlsx')
        print("Processo concluído")

    finally:
        # Fecha o navegador
        driver.quit()


# Executa a função
sbPortal_PDL()
