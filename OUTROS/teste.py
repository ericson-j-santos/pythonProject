import openpyxl
import requests
from bs4 import BeautifulSoup
from googlesearch import search
import matplotlib.pyplot as plt


# Função para carregar os dados do Excel
def carregar_dados_do_excel(arquivo_excel):
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        sheet = workbook.active
        dados = []
        # Começamos a partir da segunda linha para pular o cabeçalho
        for row in sheet.iter_rows(min_row=2, values_only=True):
            arquivo, titulo, paragrafo = row
            dados.append({"arquivo": arquivo, "titulo": titulo, "paragrafo": paragrafo})
        return dados
    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")
        return []

# Função para pesquisar no Google
def pesquisar_no_google(consulta, arquivo_word):
    resultados = []
    print(f"\nRealizando pesquisa para o arquivo: {arquivo_word}")
    try:
        for resultado in search(consulta, num_results=10, lang='pt'):
            print(f"Encontrado para {arquivo_word}: {resultado}")
            resultados.append(resultado)
    except Exception as e:
        print(f"Erro ao realizar pesquisa no Google para {arquivo_word}: {e}")
    return resultados

# Função para validar os links retornados
def validar_link(link, palavras_chave):
    try:
        response = requests.get(link)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            texto = soup.get_text()
            for palavra in palavras_chave:
                if palavra.lower() in texto.lower():
                    return True
        return False
    except Exception as e:
        print(f"Erro ao acessar o link {link}: {e}")
        return False

# Função para salvar os resultados da pesquisa em um arquivo Excel
def salvar_resultados_pesquisa_no_excel(resultados_pesquisa, arquivo_excel):
    # Criando ou abrindo o arquivo Excel para salvar os resultados da pesquisa
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.append(['Arquivo', 'Link Encontrado', 'Validado'])

    sheet = workbook.active

    # Salvando os resultados da pesquisa
    for resultado in resultados_pesquisa:
        arquivo, links, validos = resultado['arquivo'], resultado['resultados'], resultado['validos']
        for link, valido in zip(links, validos):
            sheet.append([arquivo, link, "Sim" if valido else "Não"])

    # Salvando o arquivo Excel
    workbook.save(arquivo_excel)
    print(f"Resultados da pesquisa salvos no arquivo {arquivo_excel}")

# Função para realizar a pesquisa no Google com base nos dados extraídos do Excel e validar os links
def realizar_pesquisa_google_e_validacao(arquivo_excel, arquivo_resultado_excel):
    dados = carregar_dados_do_excel(arquivo_excel)
    todos_os_resultados = []

    for dado in dados:
        arquivo = dado["arquivo"]
        titulo = dado["titulo"]
        paragrafo = dado["paragrafo"]
        consulta = f'{titulo} {paragrafo}'
        palavras_chave = titulo.split() + paragrafo.split()

        # Pesquisa no Google
        resultados_pesquisa = pesquisar_no_google(consulta, arquivo)

        # Validação dos links
        resultados_validos = []
        for link in resultados_pesquisa:
            valido = validar_link(link, palavras_chave)
            resultados_validos.append(valido)

        # Salvando os resultados em uma lista
        todos_os_resultados.append({
            "arquivo": arquivo,
            "resultados": resultados_pesquisa,
            "validos": resultados_validos
        })

    # Salvar os resultados da pesquisa e da validação no arquivo Excel
    salvar_resultados_pesquisa_no_excel(todos_os_resultados, arquivo_resultado_excel)

# Função para gerar relatórios com base nos resultados validados
def gerar_relatorio(arquivo_excel):
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        sheet = workbook.active
        total_links = 0
        validos = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            total_links += 1
            if row[2] == "Sim":
                validos += 1

        # Gerando gráfico de pizza
        labels = ['Links Válidos', 'Links Inválidos']
        sizes = [validos, total_links - validos]
        plt.figure(figsize=(6, 6))
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
        plt.axis('equal')  # Assegura que o gráfico seja um círculo
        plt.title('Distribuição de Links Válidos e Inválidos')
        plt.show()

    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")

# Exemplo de uso
arquivo_excel = 'resultados_arquivos_word.xlsx'  # Dados extraídos dos arquivos Word
arquivo_resultado_excel = 'resultados_pesquisa_google_validacao.xlsx'  # Resultados da pesquisa Google validados

realizar_pesquisa_google_e_validacao(arquivo_excel, arquivo_resultado_excel)

# Gerar relatório com gráfico
gerar_relatorio(arquivo_resultado_excel)
