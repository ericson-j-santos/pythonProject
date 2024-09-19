import openpyxl
import requests
from bs4 import BeautifulSoup
import time
import os
import random

# Função para carregar os resultados do arquivo Excel
def carregar_resultados_do_excel(arquivo_excel):
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        sheet = workbook.active
        resultados = []
        # Começamos a partir da segunda linha para pular o cabeçalho
        for row in sheet.iter_rows(min_row=2, values_only=True):
            arquivo, link = row
            resultados.append({"arquivo": arquivo, "link": link})
        return resultados
    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")
        return []

# Função para baixar a página e salvar em arquivo local com retries e random User-Agent
def baixar_pagina(link, arquivo_destino, tentativas=3):
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36",
        "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36"
    ]
    proxies = {
        # Exemplo de proxies gratuitos, utilize proxies mais robustos se necessário.
        "http": "http://10.10.10.10:8000",
        "https": "http://10.10.10.10:8000"
    }

    for tentativa in range(tentativas):
        try:
            headers = {"User-Agent": random.choice(user_agents)}
            response = requests.get(link, headers=headers, timeout=15)  # Adicione proxies=proxies se necessário
            response.raise_for_status()
            # Salvar o conteúdo HTML da página em um arquivo
            with open(arquivo_destino, 'w', encoding='utf-8') as file:
                file.write(response.text)
            return arquivo_destino
        except requests.exceptions.RequestException as e:
            print(f"Tentativa {tentativa + 1} - Erro ao baixar o link {link}: {e}")
            if "429" in str(e) or "403" in str(e):
                print(f"Aguardando tempo adicional devido ao erro {e}...")
                time.sleep(random.randint(30, 60))  # Tempo aleatório de espera entre 30 e 60 segundos
            else:
                time.sleep(5)  # Espera curta para outros erros
    return None

# Função para validar se o conteúdo do arquivo HTML baixado contém as palavras-chave
def validar_conteudo_do_arquivo(arquivo, palavras_chave):
    try:
        with open(arquivo, 'r', encoding='utf-8') as file:
            texto = file.read().lower()
        return all(palavra.lower() in texto for palavra in palavras_chave)
    except Exception as e:
        print(f"Erro ao ler ou validar o arquivo {arquivo}: {e}")
        return False

# Carregar os resultados do arquivo Excel
arquivo_resultado_excel = 'resultados_pesquisa_google.xlsx'
resultados_pesquisa = carregar_resultados_do_excel(arquivo_resultado_excel)

# Inicializando contadores para os resumos
total_validos = 0
total_invalidos = 0
total_erros = 0
erro_404 = 0
erro_403 = 0
erro_429 = 0
erro_timeout = 0
erro_formato = 0

# Diretório para salvar os arquivos HTML baixados
diretorio_html = 'paginas_baixadas'

if not os.path.exists(diretorio_html):
    os.makedirs(diretorio_html)

# Arquivo para salvar logs de erros detalhados
arquivo_erros = 'erros_log.txt'

# Validar os links com base nas palavras-chave
resultados_validados = []
for resultado in resultados_pesquisa:
    arquivo = resultado['arquivo']
    link = resultado['link']

    # Verifica se o nome do arquivo contém o separador esperado
    if " - " in arquivo:
        try:
            titulo = arquivo.split(" - ")[1].replace(".docx", "")
            paragrafo = ""  # Aqui podemos ajustar para incluir o parágrafo real, se disponível
        except IndexError:
            print(f"Formato do arquivo inesperado: {arquivo}")
            erro_formato += 1
            continue

        # Lista de palavras-chave a partir do título
        palavras_chave = titulo.split()  # Use uma abordagem mais complexa conforme necessário

        # Nome do arquivo HTML para salvar o conteúdo da página
        nome_arquivo_html = os.path.join(diretorio_html, f"pagina_{resultados_pesquisa.index(resultado)}.html")

        # Baixar a página HTML
        arquivo_baixado = baixar_pagina(link, nome_arquivo_html)

        if arquivo_baixado:
            # Validar o conteúdo do arquivo HTML baixado
            resultado_validacao = validar_conteudo_do_arquivo(arquivo_baixado, palavras_chave)
        else:
            resultado_validacao = "Erro ao baixar"

        if resultado_validacao == True:
            resultados_validados.append({"arquivo": arquivo, "link": link, "valido": True})
            total_validos += 1
        elif resultado_validacao == False:
            resultados_validados.append({"arquivo": arquivo, "link": link, "valido": False})
            total_invalidos += 1
        else:
            total_erros += 1
            # Categoriza erros específicos
            if "404" in resultado_validacao:
                erro_404 += 1
            elif "403" in resultado_validacao:
                erro_403 += 1
            elif "429" in resultado_validacao:
                erro_429 += 1
            elif "Read timed out" in resultado_validacao or "Max retries exceeded" in resultado_validacao:
                erro_timeout += 1
            # Salvar erro detalhado no log
            with open(arquivo_erros, 'a', encoding='utf-8') as log:
                log.write(f"Erro ao validar arquivo: {arquivo}\nLink: {link}\nErro: {resultado_validacao}\n\n")
            resultados_validados.append(
                {"arquivo": arquivo, "link": link, "valido": "Erro", "erro": resultado_validacao})
    else:
        print(f"Arquivo com formato inesperado: {arquivo}")
        erro_formato += 1

# Exibindo os resultados validados
print("\nResultados Validados:")
for resultado in resultados_validados:
    status = "Válido" if resultado.get("valido") == True else "Inválido" if resultado.get("valido") == False else "Erro"
    print(f"Arquivo: {resultado['arquivo']}\nLink: {resultado['link']}\nStatus: {status}\n")

# Exibindo o resumo dos resultados
print("\nResumo dos Resultados:")
print(f"Total de Links Validados como Válidos: {total_validos}")
print(f"Total de Links Validados como Inválidos: {total_invalidos}")
print(f"Total de Erros de Validação: {total_erros}")
print(f" - Erros 404 (Not Found): {erro_404}")
print(f" - Erros 403 (Forbidden): {erro_403}")
print(f" - Erros 429 (Too Many Requests): {erro_429}")
print(f" - Erros de Timeout: {erro_timeout}")
print(f"Total de Arquivos com Formato Inesperado: {erro_formato}")



# retornou apenas 24 validos
# import openpyxl
# import requests
# from bs4 import BeautifulSoup
# import time
# import os
#
# # Função para carregar os resultados do arquivo Excel
# def carregar_resultados_do_excel(arquivo_excel):
#     try:
#         workbook = openpyxl.load_workbook(arquivo_excel)
#         sheet = workbook.active
#         resultados = []
#         # Começamos a partir da segunda linha para pular o cabeçalho
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             arquivo, link = row
#             resultados.append({"arquivo": arquivo, "link": link})
#         return resultados
#     except FileNotFoundError:
#         print(f"Arquivo {arquivo_excel} não encontrado.")
#         return []
#
# # Função para baixar a página e salvar em arquivo local com retries
# def baixar_pagina(link, arquivo_destino, tentativas=3):
#     headers = {
#         "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36"
#     }
#     for tentativa in range(tentativas):
#         try:
#             response = requests.get(link, headers=headers, timeout=15)
#             response.raise_for_status()
#             # Salvar o conteúdo HTML da página em um arquivo
#             with open(arquivo_destino, 'w', encoding='utf-8') as file:
#                 file.write(response.text)
#             return arquivo_destino
#         except requests.exceptions.RequestException as e:
#             print(f"Tentativa {tentativa + 1} - Erro ao baixar o link {link}: {e}")
#             time.sleep(2)  # Espera antes de tentar novamente
#     return None
#
# # Função para validar se o conteúdo do arquivo HTML baixado contém as palavras-chave
# def validar_conteudo_do_arquivo(arquivo, palavras_chave):
#     try:
#         with open(arquivo, 'r', encoding='utf-8') as file:
#             texto = file.read().lower()
#         return all(palavra.lower() in texto for palavra in palavras_chave)
#     except Exception as e:
#         print(f"Erro ao ler ou validar o arquivo {arquivo}: {e}")
#         return False
#
# # Carregar os resultados do arquivo Excel
# arquivo_resultado_excel = 'resultados_pesquisa_google.xlsx'
# resultados_pesquisa = carregar_resultados_do_excel(arquivo_resultado_excel)
#
# # Inicializando contadores para os resumos
# total_validos = 0
# total_invalidos = 0
# total_erros = 0
# erro_404 = 0
# erro_403 = 0
# erro_429 = 0
# erro_timeout = 0
# erro_formato = 0
#
# # Diretório para salvar os arquivos HTML baixados
# diretorio_html = 'paginas_baixadas'
#
# if not os.path.exists(diretorio_html):
#     os.makedirs(diretorio_html)
#
# # Arquivo para salvar logs de erros detalhados
# arquivo_erros = 'erros_log.txt'
#
# # Validar os links com base nas palavras-chave
# resultados_validados = []
# for resultado in resultados_pesquisa:
#     arquivo = resultado['arquivo']
#     link = resultado['link']
#
#     # Verifica se o nome do arquivo contém o separador esperado
#     if " - " in arquivo:
#         try:
#             titulo = arquivo.split(" - ")[1].replace(".docx", "")
#             paragrafo = ""  # Aqui podemos ajustar para incluir o parágrafo real, se disponível
#         except IndexError:
#             print(f"Formato do arquivo inesperado: {arquivo}")
#             erro_formato += 1
#             continue
#
#         # Lista de palavras-chave a partir do título
#         palavras_chave = titulo.split()  # Use uma abordagem mais complexa conforme necessário
#
#         # Nome do arquivo HTML para salvar o conteúdo da página
#         nome_arquivo_html = os.path.join(diretorio_html, f"pagina_{resultados_pesquisa.index(resultado)}.html")
#
#         # Baixar a página HTML
#         arquivo_baixado = baixar_pagina(link, nome_arquivo_html)
#
#         if arquivo_baixado:
#             # Validar o conteúdo do arquivo HTML baixado
#             resultado_validacao = validar_conteudo_do_arquivo(arquivo_baixado, palavras_chave)
#         else:
#             resultado_validacao = "Erro ao baixar"
#
#         if resultado_validacao == True:
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": True})
#             total_validos += 1
#         elif resultado_validacao == False:
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": False})
#             total_invalidos += 1
#         else:
#             total_erros += 1
#             # Categoriza erros específicos
#             if "404" in resultado_validacao:
#                 erro_404 += 1
#             elif "403" in resultado_validacao:
#                 erro_403 += 1
#             elif "429" in resultado_validacao:
#                 erro_429 += 1
#             elif "Read timed out" in resultado_validacao or "Max retries exceeded" in resultado_validacao:
#                 erro_timeout += 1
#             # Salvar erro detalhado no log
#             with open(arquivo_erros, 'a', encoding='utf-8') as log:
#                 log.write(f"Erro ao validar arquivo: {arquivo}\nLink: {link}\nErro: {resultado_validacao}\n\n")
#             resultados_validados.append(
#                 {"arquivo": arquivo, "link": link, "valido": "Erro", "erro": resultado_validacao})
#     else:
#         print(f"Arquivo com formato inesperado: {arquivo}")
#         erro_formato += 1
#
# # Exibindo os resultados validados
# print("\nResultados Validados:")
# for resultado in resultados_validados:
#     status = "Válido" if resultado.get("valido") == True else "Inválido" if resultado.get("valido") == False else "Erro"
#     print(f"Arquivo: {resultado['arquivo']}\nLink: {resultado['link']}\nStatus: {status}\n")
#
# # Exibindo o resumo dos resultados
# print("\nResumo dos Resultados:")
# print(f"Total de Links Validados como Válidos: {total_validos}")
# print(f"Total de Links Validados como Inválidos: {total_invalidos}")
# print(f"Total de Erros de Validação: {total_erros}")
# print(f" - Erros 404 (Not Found): {erro_404}")
# print(f" - Erros 403 (Forbidden): {erro_403}")
# print(f" - Erros 429 (Too Many Requests): {erro_429}")
# print(f" - Erros de Timeout: {erro_timeout}")
# print(f"Total de Arquivos com Formato Inesperado: {erro_formato}")
#
#





# deu certo o retorno, mas apenas 24 validos
# import openpyxl
# import requests
# from bs4 import BeautifulSoup
# import time
#
#
# # Função para carregar os resultados do arquivo Excel
# def carregar_resultados_do_excel(arquivo_excel):
#     try:
#         workbook = openpyxl.load_workbook(arquivo_excel)
#         sheet = workbook.active
#         resultados = []
#         # Começamos a partir da segunda linha para pular o cabeçalho
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             arquivo, link = row
#             resultados.append({"arquivo": arquivo, "link": link})
#         return resultados
#     except FileNotFoundError:
#         print(f"Arquivo {arquivo_excel} não encontrado.")
#         return []
#
#
# # Função para baixar a página e salvar em arquivo local
# def baixar_pagina(link, arquivo_destino):
#     try:
#         response = requests.get(link, timeout=10)
#         response.raise_for_status()
#         # Salvar o conteúdo HTML da página em um arquivo
#         with open(arquivo_destino, 'w', encoding='utf-8') as file:
#             file.write(response.text)
#         return arquivo_destino
#     except requests.exceptions.RequestException as e:
#         print(f"Erro ao baixar o link {link}: {e}")
#         return None
#
#
# # Função para validar se o conteúdo do arquivo HTML baixado contém as palavras-chave
# def validar_conteudo_do_arquivo(arquivo, palavras_chave):
#     try:
#         with open(arquivo, 'r', encoding='utf-8') as file:
#             texto = file.read().lower()
#         return all(palavra.lower() in texto for palavra in palavras_chave)
#     except Exception as e:
#         print(f"Erro ao ler ou validar o arquivo {arquivo}: {e}")
#         return False
#
#
# # Carregar os resultados do arquivo Excel
# arquivo_resultado_excel = 'resultados_pesquisa_google.xlsx'
# resultados_pesquisa = carregar_resultados_do_excel(arquivo_resultado_excel)
#
# # Inicializando contadores para os resumos
# total_validos = 0
# total_invalidos = 0
# total_erros = 0
# erro_404 = 0
# erro_403 = 0
# erro_429 = 0
# erro_timeout = 0
# erro_formato = 0
#
# # Diretório para salvar os arquivos HTML baixados
# diretorio_html = 'paginas_baixadas'
# import os
#
# if not os.path.exists(diretorio_html):
#     os.makedirs(diretorio_html)
#
# # Validar os links com base nas palavras-chave
# resultados_validados = []
# for resultado in resultados_pesquisa:
#     arquivo = resultado['arquivo']
#     link = resultado['link']
#
#     # Verifica se o nome do arquivo contém o separador esperado
#     if " - " in arquivo:
#         try:
#             titulo = arquivo.split(" - ")[1].replace(".docx", "")
#             paragrafo = ""  # Aqui podemos ajustar para incluir o parágrafo real, se disponível
#         except IndexError:
#             print(f"Formato do arquivo inesperado: {arquivo}")
#             erro_formato += 1
#             continue
#
#         # Lista de palavras-chave a partir do título
#         palavras_chave = titulo.split()  # Use uma abordagem mais complexa conforme necessário
#
#         # Nome do arquivo HTML para salvar o conteúdo da página
#         nome_arquivo_html = os.path.join(diretorio_html, f"pagina_{resultados_pesquisa.index(resultado)}.html")
#
#         # Baixar a página HTML
#         arquivo_baixado = baixar_pagina(link, nome_arquivo_html)
#
#         if arquivo_baixado:
#             # Validar o conteúdo do arquivo HTML baixado
#             resultado_validacao = validar_conteudo_do_arquivo(arquivo_baixado, palavras_chave)
#         else:
#             resultado_validacao = "Erro ao baixar"
#
#         if resultado_validacao == True:
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": True})
#             total_validos += 1
#         elif resultado_validacao == False:
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": False})
#             total_invalidos += 1
#         else:
#             total_erros += 1
#             # Categoriza erros específicos
#             if "404" in resultado_validacao:
#                 erro_404 += 1
#             elif "403" in resultado_validacao:
#                 erro_403 += 1
#             elif "429" in resultado_validacao:
#                 erro_429 += 1
#             elif "Read timed out" in resultado_validacao or "Max retries exceeded" in resultado_validacao:
#                 erro_timeout += 1
#             resultados_validados.append(
#                 {"arquivo": arquivo, "link": link, "valido": "Erro", "erro": resultado_validacao})
#     else:
#         print(f"Arquivo com formato inesperado: {arquivo}")
#         erro_formato += 1
#
# # Exibindo os resultados validados
# print("\nResultados Validados:")
# for resultado in resultados_validados:
#     status = "Válido" if resultado.get("valido") == True else "Inválido" if resultado.get("valido") == False else "Erro"
#     print(f"Arquivo: {resultado['arquivo']}\nLink: {resultado['link']}\nStatus: {status}\n")
#
# # Exibindo o resumo dos resultados
# print("\nResumo dos Resultados:")
# print(f"Total de Links Validados como Válidos: {total_validos}")
# print(f"Total de Links Validados como Inválidos: {total_invalidos}")
# print(f"Total de Erros de Validação: {total_erros}")
# print(f" - Erros 404 (Not Found): {erro_404}")
# print(f" - Erros 403 (Forbidden): {erro_403}")
# print(f" - Erros 429 (Too Many Requests): {erro_429}")
# print(f" - Erros de Timeout: {erro_timeout}")
# print(f"Total de Arquivos com Formato Inesperado: {erro_formato}")
#






# deu certo com retornos, mas só 24 validos
# import openpyxl
# import requests
# from bs4 import BeautifulSoup
# import time
#
#
# # Função para carregar os resultados do arquivo Excel
# def carregar_resultados_do_excel(arquivo_excel):
#     try:
#         workbook = openpyxl.load_workbook(arquivo_excel)
#         sheet = workbook.active
#         resultados = []
#         # Começamos a partir da segunda linha para pular o cabeçalho
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             arquivo, link = row
#             resultados.append({"arquivo": arquivo, "link": link})
#         return resultados
#     except FileNotFoundError:
#         print(f"Arquivo {arquivo_excel} não encontrado.")
#         return []
#
#
# # Função para validar se o link contém as palavras-chave do título e parágrafo
# def validar_conteudo_do_link(link, palavras_chave):
#     try:
#         response = requests.get(link, timeout=10)
#         response.raise_for_status()
#         soup = BeautifulSoup(response.text, 'html.parser')
#         texto = soup.get_text().lower()
#         return all(palavra.lower() in texto for palavra in palavras_chave)
#     except requests.exceptions.RequestException as e:
#         print(f"Erro ao acessar ou validar o link {link}: {e}")
#         # Retorna uma mensagem de erro específica para categorizar
#         return str(e)
#
#
# # Carregar os resultados do arquivo Excel
# arquivo_resultado_excel = 'resultados_pesquisa_google.xlsx'
# resultados_pesquisa = carregar_resultados_do_excel(arquivo_resultado_excel)
#
# # Inicializando contadores para os resumos
# total_validos = 0
# total_invalidos = 0
# total_erros = 0
# erro_404 = 0
# erro_403 = 0
# erro_429 = 0
# erro_timeout = 0
# erro_formato = 0
#
# # Validar os links com base nas palavras-chave
# resultados_validados = []
# for resultado in resultados_pesquisa:
#     arquivo = resultado['arquivo']
#     link = resultado['link']
#
#     # Verifica se o nome do arquivo contém o separador esperado
#     if " - " in arquivo:
#         try:
#             titulo = arquivo.split(" - ")[1].replace(".docx", "")
#             paragrafo = ""  # Aqui podemos ajustar para incluir o parágrafo real, se disponível
#         except IndexError:
#             print(f"Formato do arquivo inesperado: {arquivo}")
#             erro_formato += 1
#             continue
#
#         # Lista de palavras-chave a partir do título
#         palavras_chave = titulo.split()  # Use uma abordagem mais complexa conforme necessário
#
#         # Adicionando um atraso entre as requisições para evitar bloqueio
#         time.sleep(5)  # Atraso de 5 segundos entre requisições
#
#         resultado_validacao = validar_conteudo_do_link(link, palavras_chave)
#
#         if resultado_validacao == True:
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": True})
#             total_validos += 1
#         elif resultado_validacao == False:
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": False})
#             total_invalidos += 1
#         else:
#             total_erros += 1
#             # Categoriza erros específicos
#             if "404" in resultado_validacao:
#                 erro_404 += 1
#             elif "403" in resultado_validacao:
#                 erro_403 += 1
#             elif "429" in resultado_validacao:
#                 erro_429 += 1
#             elif "Read timed out" in resultado_validacao or "Max retries exceeded" in resultado_validacao:
#                 erro_timeout += 1
#             resultados_validados.append(
#                 {"arquivo": arquivo, "link": link, "valido": "Erro", "erro": resultado_validacao})
#     else:
#         print(f"Arquivo com formato inesperado: {arquivo}")
#         erro_formato += 1
#
# # Exibindo os resultados validados
# print("\nResultados Validados:")
# for resultado in resultados_validados:
#     status = "Válido" if resultado.get("valido") == True else "Inválido" if resultado.get("valido") == False else "Erro"
#     print(f"Arquivo: {resultado['arquivo']}\nLink: {resultado['link']}\nStatus: {status}\n")
#
# # Exibindo o resumo dos resultados
# print("\nResumo dos Resultados:")
# print(f"Total de Links Validados como Válidos: {total_validos}")
# print(f"Total de Links Validados como Inválidos: {total_invalidos}")
# print(f"Total de Erros de Validação: {total_erros}")
# print(f" - Erros 404 (Not Found): {erro_404}")
# print(f" - Erros 403 (Forbidden): {erro_403}")
# print(f" - Erros 429 (Too Many Requests): {erro_429}")
# print(f" - Erros de Timeout: {erro_timeout}")
# print(f"Total de Arquivos com Formato Inesperado: {erro_formato}")

# codigo deu retorno ...
# import openpyxl
# import requests
# from bs4 import BeautifulSoup
# import time
#
#
# # Função para carregar os resultados do arquivo Excel
# def carregar_resultados_do_excel(arquivo_excel):
#     try:
#         workbook = openpyxl.load_workbook(arquivo_excel)
#         sheet = workbook.active
#         resultados = []
#         # Começamos a partir da segunda linha para pular o cabeçalho
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             arquivo, link = row
#             resultados.append({"arquivo": arquivo, "link": link})
#         return resultados
#     except FileNotFoundError:
#         print(f"Arquivo {arquivo_excel} não encontrado.")
#         return []
#
#
# # Função para validar se o link contém as palavras-chave do título e parágrafo
# def validar_conteudo_do_link(link, palavras_chave):
#     try:
#         response = requests.get(link, timeout=10)
#         response.raise_for_status()
#         soup = BeautifulSoup(response.text, 'html.parser')
#         texto = soup.get_text().lower()
#         return all(palavra.lower() in texto for palavra in palavras_chave)
#     except requests.exceptions.RequestException as e:
#         print(f"Erro ao acessar ou validar o link {link}: {e}")
#         return False
#
#
# # Carregar os resultados do arquivo Excel
# arquivo_resultado_excel = 'resultados_pesquisa_google.xlsx'
# resultados_pesquisa = carregar_resultados_do_excel(arquivo_resultado_excel)
#
# # Validar os links com base nas palavras-chave
# resultados_validados = []
# for resultado in resultados_pesquisa:
#     arquivo = resultado['arquivo']
#     link = resultado['link']
#
#     # Verifica se o nome do arquivo contém o separador esperado
#     if " - " in arquivo:
#         try:
#             titulo = arquivo.split(" - ")[1].replace(".docx", "")
#             paragrafo = ""  # Aqui podemos ajustar para incluir o parágrafo real, se disponível
#         except IndexError:
#             print(f"Formato do arquivo inesperado: {arquivo}")
#             continue
#
#         # Lista de palavras-chave a partir do título
#         palavras_chave = titulo.split()  # Use uma abordagem mais complexa conforme necessário
#
#         # Adicionando um atraso entre as requisições para evitar bloqueio
#         time.sleep(1)  # Atraso de 1 segundo entre requisições
#
#         if validar_conteudo_do_link(link, palavras_chave):
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": True})
#         else:
#             resultados_validados.append({"arquivo": arquivo, "link": link, "valido": False})
#     else:
#         print(f"Arquivo com formato inesperado: {arquivo}")
#
# # Exibindo os resultados validados
# print("\nResultados Validados:")
# for resultado in resultados_validados:
#     status = "Válido" if resultado["valido"] else "Inválido"
#     print(f"Arquivo: {resultado['arquivo']}\nLink: {resultado['link']}\nStatus: {status}\n")




# import openpyxl
# import requests
# from bs4 import BeautifulSoup
#
#
# # Função para carregar os resultados do arquivo Excel
# def carregar_resultados_do_excel(arquivo_excel):
#     try:
#         workbook = openpyxl.load_workbook(arquivo_excel)
#         sheet = workbook.active
#         resultados = []
#         # Começamos a partir da segunda linha para pular o cabeçalho
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             arquivo, link = row
#             resultados.append({"arquivo": arquivo, "link": link})
#         return resultados
#     except FileNotFoundError:
#         print(f"Arquivo {arquivo_excel} não encontrado.")
#         return []
#
#
# # Função para validar se o link contém as palavras-chave do título e parágrafo
# def validar_conteudo_do_link(link, palavras_chave):
#     try:
#         response = requests.get(link, timeout=5)
#         response.raise_for_status()
#         soup = BeautifulSoup(response.text, 'html.parser')
#         texto = soup.get_text().lower()
#         return all(palavra.lower() in texto for palavra in palavras_chave)
#     except Exception as e:
#         print(f"Erro ao acessar ou validar o link {link}: {e}")
#         return False
#
#
# # Carregar os resultados do arquivo Excel
# arquivo_resultado_excel = 'resultados_pesquisa_google.xlsx'
# resultados_pesquisa = carregar_resultados_do_excel(arquivo_resultado_excel)
#
# # Validar os links com base nas palavras-chave
# resultados_validados = []
# for resultado in resultados_pesquisa:
#     arquivo = resultado['arquivo']
#     link = resultado['link']
#     # Extraindo título (supõe que o nome do arquivo segue um padrão)
#     titulo, paragrafo = arquivo.split(" - ")[1].replace(".docx", ""), ""  # Extraindo título para simular a separação
#
#     # Lista de palavras-chave a partir do título
#     palavras_chave = titulo.split()  # Use uma abordagem mais complexa conforme necessário
#     if validar_conteudo_do_link(link, palavras_chave):
#         resultados_validados.append({"arquivo": arquivo, "link": link, "valido": True})
#     else:
#         resultados_validados.append({"arquivo": arquivo, "link": link, "valido": False})
#
# # Exibindo os resultados validados
# print("\nResultados Validados:")
# for resultado in resultados_validados:
#     status = "Válido" if resultado["valido"] else "Inválido"
#     print(f"Arquivo: {resultado['arquivo']}\nLink: {resultado['link']}\nStatus: {status}\n")
