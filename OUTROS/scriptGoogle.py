import openpyxl
from googlesearch import search

# Função para carregar os dados do Excel
def carregar_dados_do_excel(arquivo_excel):
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        sheet = workbook.active
        dados = []
        # Começamos a partir da segunda linha para pular o cabeçalho
        for row in sheet.iter_rows(min_row=2, values_only=True):
            arquivo, titulo, paragrafo = row
            dados.append({"arquivo": arquivo, "titulo": titulo, "paragrafo": paragrafo})
        return dados
    except FileNotFoundError:
        print(f"Arquivo {arquivo_excel} não encontrado.")
        return []

# Função para pesquisar no Google
def pesquisar_no_google(consulta, arquivo_word):
    resultados = []
    print(f"\nRealizando pesquisa para o arquivo: {arquivo_word}")
    try:
        for resultado in search(consulta, num_results=10, lang='pt'):
            print(f"Encontrado para {arquivo_word}: {resultado}")
            resultados.append(resultado)
    except Exception as e:
        print(f"Erro ao realizar pesquisa no Google para {arquivo_word}: {e}")
    return resultados

# Função para salvar os resultados da pesquisa em um arquivo Excel
def salvar_resultados_pesquisa_no_excel(resultados_pesquisa, arquivo_excel):
    # Criando ou abrindo o arquivo Excel para salvar os resultados da pesquisa
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.append(['Arquivo', 'Link Encontrado'])

    sheet = workbook.active

    # Salvando os resultados da pesquisa
    for resultado in resultados_pesquisa:
        arquivo, links = resultado['arquivo'], resultado['resultados']
        for link in links:
            sheet.append([arquivo, link])

    # Salvando o arquivo Excel
    workbook.save(arquivo_excel)
    print(f"Resultados da pesquisa salvos no arquivo {arquivo_excel}")

# Função para realizar a pesquisa no Google com base nos dados extraídos do Excel
def realizar_pesquisa_google_a_partir_do_excel(arquivo_excel, arquivo_resultado_excel):
    dados = carregar_dados_do_excel(arquivo_excel)
    todos_os_resultados = []

    for dado in dados:
        arquivo = dado["arquivo"]
        consulta = f'{dado["titulo"]} {dado["paragrafo"]}'
        resultados_pesquisa = pesquisar_no_google(consulta, arquivo)

        # Salvando os resultados em uma lista
        todos_os_resultados.append({
            "arquivo": arquivo,
            "resultados": resultados_pesquisa
        })

    # Salvar os resultados da pesquisa no arquivo Excel
    salvar_resultados_pesquisa_no_excel(todos_os_resultados, arquivo_resultado_excel)

# Exemplo de uso
arquivo_excel = 'resultados_arquivos_word.xlsx'  # Dados extraídos dos arquivos Word
arquivo_resultado_excel = 'resultados_pesquisa_google.xlsx'  # Resultados da pesquisa Google
realizar_pesquisa_google_a_partir_do_excel(arquivo_excel, arquivo_resultado_excel)





# from googlesearch import search
#
# def pesquisar_no_google(consulta):
#     resultados = []
#     try:
#         for resultado in search(consulta, num_results=10, lang='pt'):
#             print(f"Encontrado: {resultado}")
#             resultados.append(resultado)
#     except Exception as e:
#         print(f"Erro ao realizar pesquisa no Google: {e}")
#     return resultados
#
# # Testando com um exemplo
# consulta = "FECAP Connect oferece networking e workshop para empresas"
# resultados_pesquisa = pesquisar_no_google(consulta)
#
# if resultados_pesquisa:
#     print("Resultados da pesquisa:")
#     for resultado in resultados_pesquisa:
#         print(resultado)
# else:
#     print("Nenhum resultado encontrado.")
